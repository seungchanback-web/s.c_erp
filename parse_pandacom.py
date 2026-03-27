#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
parse_pandacom.py
Parses 10 Pandacom (팬다콤) Excel trade documents and saves to JSON.

Format A: Monthly trade documents (거래명세서/청구서/명세표) - 9 files
  3월-5월, 7월-10월: date=C1, item=C2, spec=C7, qty=C9, price=C10, amount=C11
  6월:               date=C2, item=C3, spec=C8, qty=C10, price=C11, amount=C12
  11월:              product header in row 10 C2, items in C3, qty=C9, price=C11, amount=C12

Format B: Quotation (견적서) - 1 file
  item=C2, spec=C6+C7, qty=C9, unit=C10, price=C11, amount=C12

Header row is always row 9. Data starts at row 10.
Korean header keywords used for layout detection:
  일자 = date column
  내역 = item/process (3월-10월)
  품명 = product name (11월)
  규격 or 작업 = spec column
  수량 = qty
  단가 = unit price
  금액 = amount (first occurrence)
  세액 = tax amount
"""

import openpyxl
import json
import re
import os
import datetime

# ---------------------------------------------------------------------------
# File list
# ---------------------------------------------------------------------------
TRADE_FILES = [
    ('2025-03', r'C:\Users\bungb\Downloads\팬다콤_바른컴퍼니-3월분.XLSX'),
    ('2025-04', r'C:\Users\bungb\Downloads\팬다콤_바른컴퍼니-4월분.XLSX'),
    ('2025-05', r'C:\Users\bungb\Downloads\팬다콤_바른컴퍼니-5월분.XLSX'),
    ('2025-06', r'C:\Users\bungb\Downloads\팬다콤_바른컴퍼니-6월분.XLSX'),
    ('2025-07', r'C:\Users\bungb\Downloads\팬다콤_바른컴퍼니-7월분.XLSX'),
    ('2025-08', r'C:\Users\bungb\Downloads\팬다콤_바른컴퍼니-8월분명세서.XLSX'),
    ('2025-09', r'C:\Users\bungb\Downloads\팬다콤_바른컴퍼니-9월분.XLSX'),
    ('2025-10', r'C:\Users\bungb\Downloads\팬다콤_바른컴퍼니-10월분.XLSX'),
    ('2025-11', r'C:\Users\bungb\Downloads\팬다콤_바른컴퍼니-11월 명세표.xlsx'),
]

QUOTE_FILE = r'C:\Users\bungb\Downloads\팬다콤_더 기프트 4단 리플릿 견적서.XLSX'

# Korean keywords as actual Unicode strings
VENDOR_NAME = '팬다콤'
KW_EOLJA = '일자'      # date column header
KW_NAEYEOK = '내역'   # item/process column header
KW_PUMYEONG = '품명'  # product name column header
KW_GYUGYEOK = '규격'  # spec column header
KW_JAGEOB = '작업'    # work/spec column header (11월)
KW_SURYANG = '수량'   # qty column header
KW_DANGA = '단가'     # unit price column header
KW_GEUMMAEK = '금액'  # amount column header
KW_SEAEK = '세액'     # tax column header

# Summary row keywords to skip
SKIP_KW = ['합계', '소계', '총액', '선급금', '결제', '차감', '잔액', '합   계', '합  계']


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_str(val):
    """Strip and clean a cell value to string."""
    if val is None:
        return ''
    s = str(val).strip()
    s = re.sub(r'_x[0-9A-Fa-f]{4}_', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def clean_num(val):
    """Convert a cell value to float, stripping units and commas."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = re.sub(r'_x[0-9A-Fa-f]{4}_', '', s)
    # Keep only digits, dot, minus
    s = re.sub(r'[^\d.\-]', '', s)
    s = s.strip()
    if not s or s in ('-', '.'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def date_to_str(val, year_month):
    """Convert a date cell value to YYYY-MM-DD string."""
    if isinstance(val, datetime.datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, datetime.date):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)) and 44000 < val < 47000:
        try:
            d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(val))
            return d.strftime('%Y-%m-%d')
        except Exception:
            pass
    s = clean_str(val)
    if not s:
        return None
    year = int(year_month[:4])
    # MM/DD patterns like ' 3/24', '10/22', ' 9/ 3', '6/30'
    m = re.match(r'^\s*(\d{1,2})\s*/\s*(\d{1,2})\s*$', s)
    if m:
        d_month = int(m.group(1))
        d_day = int(m.group(2))
        return f'{year}-{d_month:02d}-{d_day:02d}'
    # Pattern like '2025-9.30.'
    m2 = re.match(r'(\d{4})-(\d{1,2})\.(\d{1,2})\.?', s)
    if m2:
        return f'{m2.group(1)}-{int(m2.group(2)):02d}-{int(m2.group(3)):02d}'
    return None


def is_date_value(val):
    """Return True if cell value looks like a date."""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return True
    if val is None:
        return False
    if isinstance(val, (int, float)) and 44000 < val < 47000:
        return True
    s = clean_str(val)
    if not s:
        return False
    if re.match(r'^\s*\d{1,2}\s*/\s*\d{1,2}\s*$', s):
        return True
    if re.match(r'\d{4}-\d{1,2}\.\d{1,2}', s):
        return True
    return False


def is_summary_row(text):
    """Return True if text looks like a summary/total row."""
    if not text:
        return False
    t = text.replace(' ', '')
    return any(kw.replace(' ', '') in t for kw in SKIP_KW)


# ---------------------------------------------------------------------------
# Auto-detect column layout from row 9 (header row)
# ---------------------------------------------------------------------------

def detect_layout(ws):
    """
    Scan row 9 for header keywords to determine column layout.
    Returns dict with 1-indexed column numbers and 'format' key.
    """
    max_col = ws.max_column
    layout = {
        'date_col': None,
        'item_col': None,
        'spec_col': None,
        'qty_col': None,
        'price_col': None,
        'amount_col': None,
        'tax_col': None,
        'format': 'standard',
    }

    for col in range(1, max_col + 1):
        val = ws.cell(row=9, column=col).value
        if val is None:
            continue
        s = str(val).replace(' ', '')
        if KW_EOLJA in s:
            layout['date_col'] = col
        if KW_NAEYEOK in s or KW_PUMYEONG in s:
            layout['item_col'] = col
        if KW_GYUGYEOK in s or KW_JAGEOB in s:
            layout['spec_col'] = col
        if KW_SURYANG in s:
            layout['qty_col'] = col
        if KW_DANGA in s:
            layout['price_col'] = col
        if KW_GEUMMAEK in s and not layout['amount_col']:
            layout['amount_col'] = col
        if KW_SEAEK in s:
            layout['tax_col'] = col

    # Determine format variant
    # 11월: date=C2, item=C3 (품명), and row 10 C2 is a long product header string (not a date)
    if layout['date_col'] == 2 and layout['item_col'] == 3:
        r10_c2 = ws.cell(row=10, column=2).value
        if r10_c2 is not None and not is_date_value(r10_c2):
            layout['format'] = '11월'
        else:
            layout['format'] = '6월'

    return layout


# ---------------------------------------------------------------------------
# Parse standard monthly trade files (Format A)
# ---------------------------------------------------------------------------

def parse_trade_file(year_month, path):
    """Parse a monthly trade document. Returns list of record dicts."""
    records = []
    fname = os.path.basename(path)
    print(f"\nParsing {year_month}: {fname}")

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  ERROR opening file: {e}")
        return records

    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # Get total amount from row 8
    total_amount = None
    for col in range(1, max_col + 1):
        val = ws.cell(row=8, column=col).value
        if isinstance(val, (int, float)) and val > 0:
            total_amount = float(val)
            break

    layout = detect_layout(ws)
    dc = layout['date_col']
    ic = layout['item_col']
    sc = layout['spec_col']
    qc = layout['qty_col']
    pc = layout['price_col']
    ac = layout['amount_col']
    fmt = layout['format']

    print(f"  Layout: date=C{dc}, item=C{ic}, spec=C{sc}, qty=C{qc}, price=C{pc}, amount=C{ac} [{fmt}]")
    if total_amount:
        print(f"  Total amount in file: {total_amount:,.0f}")
    else:
        print(f"  Total amount: None")

    # -----------------------------------------------------------------------
    # 11월 format: row 10 is a product header, items from row 11
    # -----------------------------------------------------------------------
    if fmt == '11월':
        current_product = None

        for row_idx in range(10, max_row + 1):
            c2_val = clean_str(ws.cell(row=row_idx, column=2).value)
            c3_val = clean_str(ws.cell(row=row_idx, column=ic).value) if ic else ''
            spec_a = clean_str(ws.cell(row=row_idx, column=sc).value) if sc else ''
            # Also try next column for secondary spec
            spec_b = clean_str(ws.cell(row=row_idx, column=(sc + 1) if sc else 8).value)
            qty_raw = ws.cell(row=row_idx, column=qc).value if qc else None
            unit_raw = ws.cell(row=row_idx, column=(qc + 1) if qc else 10).value
            price_raw = ws.cell(row=row_idx, column=pc).value if pc else None
            amount_raw = ws.cell(row=row_idx, column=ac).value if ac else None

            qty = clean_num(qty_raw)
            price = clean_num(price_raw)
            amount = clean_num(amount_raw)

            # Product header detection: C2 has content, not a date, C3 is empty, no price
            if c2_val and not is_date_value(ws.cell(row=row_idx, column=2).value):
                if len(c2_val) > 5 and not c3_val and price is None and amount is None:
                    current_product = c2_val
                    print(f"    Product: {current_product[:60]}")
                    continue

            # Skip empty rows
            if not c3_val and not c2_val and qty is None and price is None:
                continue

            process_type = c3_val if c3_val else c2_val
            if not process_type:
                continue

            # Skip summary rows
            if is_summary_row(process_type):
                continue

            # Skip rows with no qty, price, or amount (notes/headers)
            if qty is None and price is None and amount is None:
                continue

            spec_parts = [s for s in [spec_a, spec_b] if s]
            spec = ' '.join(spec_parts)
            unit = clean_str(unit_raw)

            record = {
                'doc_type': 'trade',
                'vendor': VENDOR_NAME,
                'year_month': year_month,
                'date': None,
                'product_name': current_product or '',
                'process_type': process_type,
                'spec': spec,
                'qty': qty,
                'unit': unit,
                'unit_price': price,
                'amount': amount,
                'source_file': fname,
            }
            records.append(record)

        print(f"  Records: {len(records)}")
        return records

    # -----------------------------------------------------------------------
    # Standard format (3월-10월, 6월)
    # -----------------------------------------------------------------------
    current_date = None
    current_product = None

    for row_idx in range(10, max_row + 1):
        date_raw = ws.cell(row=row_idx, column=dc).value if dc else None
        item_raw = ws.cell(row=row_idx, column=ic).value if ic else None
        spec_raw = ws.cell(row=row_idx, column=sc).value if sc else None
        qty_raw = ws.cell(row=row_idx, column=qc).value if qc else None
        price_raw = ws.cell(row=row_idx, column=pc).value if pc else None
        amount_raw = ws.cell(row=row_idx, column=ac).value if ac else None

        item_str = clean_str(item_raw)
        spec_str = clean_str(spec_raw)
        qty = clean_num(qty_raw)
        price = clean_num(price_raw)
        amount = clean_num(amount_raw)

        # Skip empty rows
        if not item_str and qty is None and price is None and amount is None:
            continue

        # Skip summary rows
        if is_summary_row(item_str):
            continue

        # Check if this row starts a new product section
        if is_date_value(date_raw) and item_str:
            date_str = date_to_str(date_raw, year_month)
            current_date = date_str
            current_product = item_str
            price_raw_str = clean_str(price_raw)
            is_header_only = (price is None or price_raw_str in ('', ' '))

            if is_header_only:
                print(f"    Product: [{current_date}] {current_product[:50]} qty={qty}")
                continue
            else:
                # Row has both date and price => emit as a work item too
                records.append({
                    'doc_type': 'trade',
                    'vendor': VENDOR_NAME,
                    'year_month': year_month,
                    'date': current_date,
                    'product_name': current_product,
                    'process_type': current_product,
                    'spec': spec_str,
                    'qty': qty,
                    'unit': '',
                    'unit_price': price,
                    'amount': amount,
                    'source_file': fname,
                })
                continue

        # Non-date rows: process items for current product
        if not item_str:
            continue

        records.append({
            'doc_type': 'trade',
            'vendor': VENDOR_NAME,
            'year_month': year_month,
            'date': current_date,
            'product_name': current_product or '',
            'process_type': item_str,
            'spec': spec_str,
            'qty': qty,
            'unit': '',
            'unit_price': price,
            'amount': amount,
            'source_file': fname,
        })

    print(f"  Records: {len(records)}")
    return records


# ---------------------------------------------------------------------------
# Parse quotation file (Format B)
# ---------------------------------------------------------------------------

def parse_quote_file(path):
    """Parse the quotation file. Returns list of record dicts."""
    records = []
    fname = os.path.basename(path)
    print(f"\nParsing quotation: {fname}")

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  ERROR opening file: {e}")
        return records

    ws = wb.active
    max_row = ws.max_row

    # Row 10: product header
    product_date_raw = ws.cell(row=10, column=1).value
    product_name_raw = ws.cell(row=10, column=2).value
    product_date = date_to_str(product_date_raw, '2026-01') if product_date_raw else '2026-01-27'
    product_name = clean_str(product_name_raw)
    print(f"  Product: {product_name}, date={product_date}")

    # Items from row 12: item=C2, spec=C6+C7, qty=C9, unit=C10, price=C11, amount=C12
    for row_idx in range(12, max_row + 1):
        item_raw = ws.cell(row=row_idx, column=2).value
        spec1_raw = ws.cell(row=row_idx, column=6).value
        spec2_raw = ws.cell(row=row_idx, column=7).value
        qty_raw = ws.cell(row=row_idx, column=9).value
        unit_raw = ws.cell(row=row_idx, column=10).value
        price_raw = ws.cell(row=row_idx, column=11).value
        amount_raw = ws.cell(row=row_idx, column=12).value

        item_str = clean_str(item_raw)
        if not item_str:
            continue
        if is_summary_row(item_str):
            continue

        spec_parts = [clean_str(spec1_raw), clean_str(spec2_raw)]
        spec = ' '.join(s for s in spec_parts if s)

        records.append({
            'doc_type': 'quotation',
            'vendor': VENDOR_NAME,
            'year_month': '2026-01',
            'date': product_date,
            'product_name': product_name,
            'process_type': item_str,
            'spec': spec,
            'qty': clean_num(qty_raw),
            'unit': clean_str(unit_raw),
            'unit_price': clean_num(price_raw),
            'amount': clean_num(amount_raw),
            'source_file': fname,
        })

    print(f"  Records: {len(records)}")
    return records


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    all_records = []

    for year_month, path in TRADE_FILES:
        records = parse_trade_file(year_month, path)
        all_records.extend(records)

    quote_records = parse_quote_file(QUOTE_FILE)
    all_records.extend(quote_records)

    output_path = r'C:\barunson\pandacom_data.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2, default=str)

    # Summary stats
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Total records: {len(all_records)}")

    by_type = {}
    by_month = {}
    by_process = {}
    total_amounts = {}

    for r in all_records:
        by_type[r['doc_type']] = by_type.get(r['doc_type'], 0) + 1
        by_month[r['year_month']] = by_month.get(r['year_month'], 0) + 1
        proc = r['process_type']
        by_process[proc] = by_process.get(proc, 0) + 1
        if r.get('amount') and r['amount'] and r['amount'] > 0:
            ym = r['year_month']
            total_amounts[ym] = total_amounts.get(ym, 0) + r['amount']

    print(f"\nBy doc type:")
    for k, v in sorted(by_type.items()):
        print(f"  {k}: {v}")

    print(f"\nBy month:")
    for k, v in sorted(by_month.items()):
        amt = total_amounts.get(k, 0)
        print(f"  {k}: {v} records, amount sum={amt:,.0f}")

    print(f"\nTop process types (top 25):")
    top_procs = sorted(by_process.items(), key=lambda x: -x[1])[:25]
    for k, v in top_procs:
        print(f"  [{v:3d}] {k[:60]}")

    products = sorted(set(r['product_name'] for r in all_records if r['product_name']))
    print(f"\nUnique products ({len(products)}):")
    for p in products:
        print(f"  - {p[:80]}")

    print(f"\nSample records (first 5 with amounts):")
    count = 0
    for r in all_records:
        if r['amount']:
            print(f"  {r['year_month']} | {r['product_name'][:25]} | {r['process_type'][:30]} | qty={r['qty']} | price={r['unit_price']} | amt={r['amount']}")
            count += 1
            if count >= 5:
                break

    print(f"\nOutput saved to: {output_path}")


if __name__ == '__main__':
    main()
