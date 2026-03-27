import openpyxl
import json
import re
import os

files = [
    ('2025-09', r'C:\Users\bungb\Downloads\코리아패키지_9월 바른컴퍼니.xlsx'),
    ('2025-10', r'C:\Users\bungb\Downloads\코리아패키지_10월 (주)바른컴퍼니.xlsx'),
    ('2025-11', r'C:\Users\bungb\Downloads\코리아패키지_11월 (주)바른컴퍼니.xlsx'),
    ('2025-12', r'C:\Users\bungb\Downloads\코리아패키지_12월 바른컴퍼니.xlsx'),
    ('2026-01', r'C:\Users\bungb\Downloads\코리아패키지_202601 바른컴퍼니.xlsx'),
]

def clean(v):
    if v is None: return ''
    return str(v).replace('\xa0', '').strip()

def parse_number(s):
    s = clean(s).replace(',', '').replace('원', '').replace('매', '').strip()
    try: return float(s)
    except: return None

all_records = []

for month_label, fpath in files:
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active

    # Find header row (날짜, 품목, 작업명, 규격, 수량, 단가, 금액)
    header_row = 9  # row 9 is header in all files

    current_date = ''
    current_product = ''
    current_qty = ''

    for r in range(10, ws.max_row + 1):
        c1 = clean(ws.cell(r, 1).value)  # 날짜
        c2 = clean(ws.cell(r, 2).value)  # 품목
        c3 = clean(ws.cell(r, 3).value)  # 작업명
        c4 = clean(ws.cell(r, 4).value)  # 규격
        c5 = clean(ws.cell(r, 5).value)  # 수량
        c6 = clean(ws.cell(r, 6).value)  # 단가
        c7 = clean(ws.cell(r, 7).value)  # 금액

        # Skip empty rows and summary rows
        if not c2 and not c3 and not c7:
            continue
        if '합계' in c1 or '합계' in c3 or '소계' in c3:
            continue
        if '공급가액' in c1 or '세액' in c1 or '합계' in c1:
            continue

        # Update date/product
        if c1 and re.match(r'20\d\d', c1):
            current_date = c1
        if c2 and not c2.startswith(('합계', '소계')):
            current_product = c2
            current_qty = c5

        # This is a work item (작업 행)
        if c3 and c7:
            amount = parse_number(c7)
            unit_price = parse_number(c6)
            qty_val = clean(c5)

            all_records.append({
                'month': month_label,
                'date': current_date,
                'product_code': current_product,
                'process': c3,
                'spec': c4,
                'qty': qty_val,
                'product_qty': current_qty,
                'unit_price': unit_price,
                'amount': amount
            })

    wb.close()

# Save raw data
with open('C:/barunson/korea_pkg_data.json', 'w', encoding='utf-8') as f:
    json.dump(all_records, f, ensure_ascii=False, indent=2)

# Analysis
print(f"총 거래 레코드: {len(all_records)}건")

# 1. Monthly totals
print("\n=== 월별 거래 총액 ===")
monthly = {}
for r in all_records:
    m = r['month']
    if m not in monthly: monthly[m] = 0
    if r['amount']: monthly[m] += r['amount']
for m in sorted(monthly.keys()):
    print(f"  {m}: {monthly[m]:,.0f}원")

# 2. Process types
print("\n=== 공정 유형별 집계 ===")
processes = {}
for r in all_records:
    p = r['process']
    if p not in processes: processes[p] = {'count': 0, 'total': 0}
    processes[p]['count'] += 1
    if r['amount']: processes[p]['total'] += r['amount']
for p, v in sorted(processes.items(), key=lambda x: -x[1]['total']):
    print(f"  {p}: {v['count']}건, {v['total']:,.0f}원")

# 3. Product frequency
print("\n=== 제품별 후공정 횟수 (상위 20) ===")
products = {}
for r in all_records:
    pc = r['product_code']
    if pc not in products: products[pc] = {'count': 0, 'total': 0, 'processes': set()}
    products[pc]['count'] += 1
    if r['amount']: products[pc]['total'] += r['amount']
    products[pc]['processes'].add(r['process'])
for pc, v in sorted(products.items(), key=lambda x: -x[1]['total'])[:20]:
    procs = ', '.join(sorted(v['processes']))
    print(f"  {pc}: {v['count']}건, {v['total']:,.0f}원 | 공정: {procs}")

# 4. Unit price analysis per process
print("\n=== 공정별 단가 분석 (단가 있는 항목만) ===")
price_by_process = {}
for r in all_records:
    if r['unit_price'] and r['unit_price'] > 0:
        p = r['process']
        if p not in price_by_process: price_by_process[p] = []
        price_by_process[p].append({
            'price': r['unit_price'],
            'product': r['product_code'],
            'qty': r['product_qty'],
            'month': r['month']
        })

for p in sorted(price_by_process.keys()):
    items = price_by_process[p]
    prices = [i['price'] for i in items]
    if len(set(prices)) > 1:
        print(f"\n  [{p}] 단가 변동 있음! ({len(items)}건)")
        for i in sorted(items, key=lambda x: x['price']):
            print(f"    {i['month']} | {i['product']} | 수량:{i['qty']} | 단가:{i['price']:,.0f}")
    else:
        print(f"  [{p}] 고정 단가: {prices[0]:,.0f}원 ({len(items)}건)")
